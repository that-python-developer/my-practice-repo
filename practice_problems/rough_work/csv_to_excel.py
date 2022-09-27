from os import path as os_path, listdir as os_listdir
from pandas import read_csv, pivot_table
from numpy import nan
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows


def pivot_df(src_data_df, pivot_column):
    if pivot_column:
        pivoted_table = pivot_table(
            src_data_df,
            values=['Measure Values'],
            index=['Chain Zone'],
            columns=pivot_column,
            aggfunc=[sum]
        )
        pivoted_table.columns = ['{}'.format(col[2]) for col in pivoted_table.columns.values]
        pivoted_table.replace({nan: 0}, inplace=True)
        return pivoted_table
    else:
        return src_data_df


def merge_header_and_data_df(src_header_df, src_data_df, folder_location, file_name):
    wb = Workbook()
    sheet = wb.worksheets[0]

    sheet.merge_cells('A1:D1')
    sheet['A1'].alignment = Alignment(horizontal='center')
    sheet['A1'].font = sheet['A2'].font = sheet['D2'].font = Font(bold=True)
    # title_fill = PatternFill(start_color='CCFFFF', end_color='CCFFFF', fill_type='solid')
    header_fill = PatternFill(start_color='A0A0A0', end_color='A0A0A0', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    sheet['A1'] = src_header_df['Title'][0]
    sheet['A2'] = f"{header_df['Start Week'][0]} - {header_df['End Week'][0]}"
    sheet['D2'] = f"Last Refresh: {header_df['Last Refresh'][0]}"
    # sheet['A1'].fill = sheet['A2'].fill = sheet['D2'].fill = title_fill

    r_idx = 4
    c_idx = 1
    rows = dataframe_to_rows(src_data_df,  index=False, header=True)
    for i, row in enumerate(rows, 1):
        for j, value in enumerate(row, 1):
            sheet.cell(row=r_idx, column=c_idx, value=value)
            sheet.cell(row=r_idx, column=c_idx).border = thin_border
            if r_idx == 4:
                sheet.cell(row=r_idx, column=c_idx).fill = header_fill
            c_idx += 1
        c_idx = 1
        r_idx += 1

    # red_font = Font(color='FFFFFF')
    # for cell in sheet["4:4"]:
    #     cell.font = red_font

    wb.save(os_path.join(folder_location, file_name))


if __name__ == '__main__':
    dashboard_pivot_mapping = {
        '1TerminalandStoreDetails': [],
        '2TenderModeBreakUp': ['Tendor Mode'],
        '3UPITransactionStatusBreakUp': ['Status'],
        '4NetworkBreakUp': ['Network'],
        '5TransactionModeBreakup': [],
        '6BankPenetrationBreakUp': [],
        '7TransactionNetworkBreakUp': []
    }

    csv_folder_location = 'D:\\Tableau_AutoMailer\\csv'
    excel_folder_location = 'D:\\Tableau_AutoMailer\\excel'
    header_folder_location = 'D:\\Tableau_AutoMailer\\header'
    csv_file_list = os_listdir(csv_folder_location)

    for file in csv_file_list:
        group_name = file.split('_')[-1].split('.')[0]
        header_file = f'8Heading_{group_name}.csv'
        header_df = read_csv(os_path.join(header_folder_location, header_file))

        data_df = read_csv(os_path.join(csv_folder_location, file))
        try:
            data_df.drop('blank', axis=1, inplace=True)
        except KeyError:
            pass

        pivot_columns = dashboard_pivot_mapping[file.split('_')[0]]
        data_df = pivot_df(data_df, pivot_columns)

        excel_file_name = f"{file.split('.')[0]}.xlsx"
        merge_header_and_data_df(header_df, data_df, excel_folder_location, excel_file_name)
