from os import path as os_path, listdir as os_listdir
from pandas import read_csv, pivot_table
from numpy import nan
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime


def pivot_df(src_data_df, dashboard):
    if dashboard == '1TerminalandStoreDetails':
        column_order = ['Total Stores', 'Total Terminals', '% NonTrx Terminals', 'Non Trx Terminals', 'American Gap',
                        'Axis Gap', 'HDFC Gap']
        pivoted_table = pivot_table(
            src_data_df,
            values=['Measure Values'],
            index=['Chain Zone'],
            columns=['Measure Names'],
            aggfunc=[sum]
        )
        pivoted_table.columns = ['{}'.format(col[2]) for col in pivoted_table.columns.values]
        pivoted_table = pivoted_table.reindex(column_order, axis=1)
        pivoted_table.reset_index(level=0, inplace=True)

    if dashboard == '2TenderModeBreakUp':
        pivoted_table = pivot_table(
            src_data_df,
            values=['Measure Values'],
            index=['Chain Zone'],
            columns=['Tendor Mode'],
            aggfunc=[sum]
        )
        pivoted_table.columns = ['{}'.format(col[2]) for col in pivoted_table.columns.values]
        pivoted_table.reset_index(level=0, inplace=True)

    elif dashboard == '3UPITransactionStatusBreakUp':
        pivoted_table = pivot_table(
            src_data_df,
            values=['Measure Values'],
            index=['Chain Zone'],
            columns=['Measure Names', 'Status'],
            aggfunc=[sum]
        )
        pivoted_table.columns = ['{} {}'.format(col[2], col[3]) for col in pivoted_table.columns.values]
        pivoted_table.reset_index(level=0, inplace=True)

    elif dashboard == '4NetworkBreakUp':
        pivoted_table = pivot_table(
            src_data_df,
            values=['Measure Values'],
            index=['Chain Zone'],
            columns=['Network'],
            aggfunc=[sum]
        )
        pivoted_table.columns = ['{}'.format(col[2]) for col in pivoted_table.columns.values]
        pivoted_table.reset_index(level=0, inplace=True)

    elif dashboard == '6BankPenetrationBreakUp':
        pivoted_table = pivot_table(
            src_data_df,
            values=['Measure Values'],
            index=['Chain Zone'],
            columns=['Acquirer Bank'],
            aggfunc=[sum]
        )
        pivoted_table.columns = ['{}'.format(col[2]) for col in pivoted_table.columns.values]
        pivoted_table.reset_index(level=0, inplace=True)

    pivoted_table.replace({nan: 0}, inplace=True)
    return pivoted_table


def merge_header_and_data_df(src_header_df, src_data_df, folder_location, file_name):
    wb = Workbook()
    sheet = wb.worksheets[0]

    sheet.merge_cells('A1:D1')
    sheet['A1'].alignment = Alignment(horizontal='center')
    sheet['A1'].font = sheet['A2'].font = sheet['D2'].font = Font(bold=True)
    title_fill = PatternFill(start_color='C9F2C9', end_color='C9F2C9', fill_type='solid')
    header_fill = PatternFill(start_color='A0A0A0', end_color='A0A0A0', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    sheet['A1'] = src_header_df['Title'][0]
    if 'Start Day' in header_df.columns and 'End Day' in header_df.columns:
        start_date_range = datetime.strptime(header_df['Start Day'][0], '%m/%d/%Y').strftime('%d-%b-%y')
        end_date_range = datetime.strptime(header_df['End Day'][0], '%m/%d/%Y').strftime('%d-%b-%y')
        sheet['A2'] = f"Time: {start_date_range} to {end_date_range}"
    elif 'Time Group' in header_df.columns:
        sheet['A2'] = f"Time: {header_df['Time Group'][0]}"
    last_refresh = datetime.strptime(header_df['Last Refresh'][0], '%m/%d/%Y').strftime('%d-%b-%y')
    sheet['D2'] = f"Last Refresh: {last_refresh}"
    sheet['A1'].fill = title_fill

    r_idx = 4
    c_idx = 1
    rows = dataframe_to_rows(src_data_df,  index=False, header=True)
    for i, row in enumerate(rows, 1):
        for j, value in enumerate(row, 1):
            sheet.cell(row=r_idx, column=c_idx, value=value)
            sheet.cell(row=r_idx, column=c_idx).border = thin_border
            if r_idx == 4:
                sheet.cell(row=r_idx, column=c_idx).fill = header_fill
            c_idx += 1
        c_idx = 1
        r_idx += 1

    # red_font = Font(color='FFFFFF')
    # for cell in sheet["4:4"]:
    #     cell.font = red_font

    wb.save(os_path.join(folder_location, file_name))


if __name__ == '__main__':
    csv_folder_location = 'D:\\Tableau_AutoMailer\\csv\\Merchant-Weekly-Dashboard\\crosstab'
    excel_folder_location = 'D:\\Tableau_AutoMailer\\excel\\Merchant-Weekly-Dashboard'
    header_folder_location = 'D:\\Tableau_AutoMailer\\csv\\Merchant-Weekly-Dashboard\\header'
    csv_file_list = os_listdir(csv_folder_location)

    for file in csv_file_list:
        group_name = file.split('_')[-1].split('.')[0]
        header_file = f'8Heading_{group_name}.csv'
        header_df = read_csv(os_path.join(header_folder_location, header_file))

        data_df = read_csv(os_path.join(csv_folder_location, file))
        try:
            data_df.drop('blank', axis=1, inplace=True)
        except KeyError:
            pass

        dashboard_name = file.split('_')[0]
        data_df = pivot_df(data_df, dashboard_name)

        excel_file_name = f"{file.split('.')[0]}.xlsx"
        merge_header_and_data_df(header_df, data_df, excel_folder_location, excel_file_name)
