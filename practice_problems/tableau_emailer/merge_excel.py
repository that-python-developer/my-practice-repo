from pandas import DataFrame, read_excel, ExcelWriter
from os import path as os_path, listdir as os_listdir
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from numpy import nan


if __name__ == '__main__':
    # ---------------------------------------- Configuring File Paths --------------------------------------------------

    excel_files_path = 'D:\\Tableau_AutoMailer\\excel\\Merchant-Weekly-Dashboard'
    merged_excel_files_path = 'D:\\Tableau_AutoMailer\\merged_excel'
    excel_files = os_listdir(excel_files_path)

    # ---------------------------------------- Merge Excel -------------------------------------------------------------

    excel_files_df = DataFrame(excel_files, columns=['excel_file_names'])
    excel_files_df['code'] = excel_files_df.excel_file_names.apply(lambda x: x.split('.xlsx')[0].split('_')[-1])
    excel_files_df = excel_files_df.groupby(['code'])['excel_file_names'].apply(','.join).reset_index()

    for index, group in excel_files_df.iterrows():
        merged_excel_file_name = f"{group['code']}.xlsx"

        writer = ExcelWriter(os_path.join(merged_excel_files_path, merged_excel_file_name))
        wb = Workbook()

        for filename in group['excel_file_names'].split(','):
            sheet_name = filename.split('_')[0][1:]

            df_excel = read_excel(os_path.join(excel_files_path, filename), engine='openpyxl')

            title = df_excel.columns[0]
            date_range = df_excel.iloc[0][0]
            last_refresh = df_excel.iloc[0][3]

            df_excel = df_excel[2:]
            new_header = df_excel.iloc[0]
            df_excel = df_excel[1:]
            df_excel.columns = new_header
            df_excel.reset_index(inplace=True, drop=True)

            try:
                df_excel.drop(nan, axis=1, inplace=True)
            except KeyError:
                pass

            sheet = wb.create_sheet(sheet_name)

            sheet.merge_cells('A1:D1')
            sheet['A1'].alignment = Alignment(horizontal='center')
            sheet['A1'].font = sheet['A2'].font = sheet['D2'].font = Font(bold=True)
            title_fill = PatternFill(start_color='C9F2C9', end_color='C9F2C9', fill_type='solid')
            header_fill = PatternFill(start_color='A0A0A0', end_color='A0A0A0', fill_type='solid')
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            sheet['A1'] = title
            sheet['A2'] = date_range
            sheet['D2'] = last_refresh
            sheet['A1'].fill = title_fill

            r_idx = 4
            c_idx = 1
            rows = dataframe_to_rows(df_excel, index=False, header=True)
            for i, row in enumerate(rows, 1):
                for j, value in enumerate(row, 1):
                    if r_idx == 4:
                        sheet.cell(row=r_idx, column=c_idx, value=value)
                        sheet.cell(row=r_idx, column=c_idx).fill = header_fill
                    else:
                        try:
                            if str(value)[-1] == '%':
                                value = float("%.2f" % round(float(value[:-1].replace(',', '')), 2))
                            else:
                                value = float("%.2f" % round(float(value.replace(',', '')), 2))
                            sheet.cell(row=r_idx, column=c_idx, value=value)
                        except (ValueError, AttributeError):
                            sheet.cell(row=r_idx, column=c_idx, value=value)
                    sheet.cell(row=r_idx, column=c_idx).border = thin_border
                    c_idx += 1
                c_idx = 1
                r_idx += 1
        wb.remove(wb['Sheet'])
        wb.save(os_path.join(merged_excel_files_path, merged_excel_file_name))
